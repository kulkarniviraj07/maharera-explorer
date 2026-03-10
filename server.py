"""
Property Index Project Explorer — Server
===================================
Set your file paths in config.py, then run:
    python server.py
"""

from flask import Flask, jsonify, session, request, send_from_directory, abort
import openpyxl, json, os, math, hashlib, secrets
from functools import wraps
from config import EXCEL_PATH, GRAPHS_DIR, SECRET_KEY, PORT

app = Flask(__name__, static_folder="static")
app.secret_key = SECRET_KEY

# ─── Users store (users.json next to server.py) ──────────────────

USERS_FILE = os.path.join(os.path.dirname(__file__), "users.json")

def load_users():
    if not os.path.exists(USERS_FILE):
        return {}
    with open(USERS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_users(users):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, indent=2)

def hash_password(password: str, salt: str = None):
    if salt is None:
        salt = secrets.token_hex(16)
    pw_hash = hashlib.sha256((salt + password).encode()).hexdigest()
    return pw_hash, salt

# ─── Auth helpers ────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            abort(401)
        return f(*args, **kwargs)
    return decorated

def clean(val):
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    return str(val).strip()

# ─── Data loading ─────────────────────────────────────────────────

def load_projects():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    headers = [clean(cell.value) for cell in ws[1]]
    projects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        p = {h: clean(v) for h, v in zip(headers, row)}
        projects.append(p)
    return projects

def load_graph(project_number):
    path = os.path.join(GRAPHS_DIR, f"project_{project_number}.json")
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def available_graph_numbers():
    if not os.path.isdir(GRAPHS_DIR):
        return set()
    nums = set()
    for fname in os.listdir(GRAPHS_DIR):
        if fname.startswith("project_") and fname.endswith(".json"):
            try:
                nums.add(int(fname.replace("project_", "").replace(".json", "")))
            except ValueError:
                pass
    return nums

# ─── Auth routes ─────────────────────────────────────────────────

@app.route("/api/signup", methods=["POST"])
def signup():
    data = request.get_json()
    name     = (data.get("name") or "").strip()
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""

    if not name or not username or not password:
        return jsonify({"ok": False, "error": "All fields are required."}), 400
    if len(username) < 3:
        return jsonify({"ok": False, "error": "Username must be at least 3 characters."}), 400
    if len(password) < 6:
        return jsonify({"ok": False, "error": "Password must be at least 6 characters."}), 400

    users = load_users()
    if username in users:
        return jsonify({"ok": False, "error": "Username already taken. Please choose another."}), 409

    pw_hash, salt = hash_password(password)
    users[username] = {"name": name, "hash": pw_hash, "salt": salt}
    save_users(users)

    session["logged_in"] = True
    session["username"]  = username
    session["name"]      = name
    return jsonify({"ok": True, "name": name})

@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json()
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""

    users = load_users()
    user  = users.get(username)
    if not user:
        return jsonify({"ok": False, "error": "Invalid username or password."}), 401

    pw_hash, _ = hash_password(password, user["salt"])
    if pw_hash != user["hash"]:
        return jsonify({"ok": False, "error": "Invalid username or password."}), 401

    session["logged_in"] = True
    session["username"]  = username
    session["name"]      = user.get("name", username)
    return jsonify({"ok": True, "name": user.get("name", username)})

@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/me")
def me():
    return jsonify({
        "logged_in": session.get("logged_in", False),
        "name":      session.get("name", ""),
        "username":  session.get("username", ""),
    })

# ─── Data routes ─────────────────────────────────────────────────

@app.route("/api/projects")
@login_required
def api_projects():
    try:
        projects = load_projects()
        graph_nums = available_graph_numbers()
        result = []
        for i, p in enumerate(projects):
            result.append({
                "index":            i,
                "project_name":     p.get("project_name", ""),
                "project_district": p.get("project_district", ""),
                "project_type":     p.get("project_type", ""),
                "has_graph":        (i + 1) in graph_nums,
            })
        return jsonify({"projects": result, "total": len(result)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/project/<int:index>")
@login_required
def api_project(index):
    try:
        projects = load_projects()
        if index < 0 or index >= len(projects):
            return jsonify({"error": "Project not found"}), 404
        project = projects[index]
        graph   = load_graph(index + 1)
        return jsonify({"project": project, "graph": graph})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ─── Serve frontend ──────────────────────────────────────────────

@app.route("/")
@app.route("/index.html")
def index():
    return send_from_directory("static", "index.html")

# ─── Run ─────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--host", default="127.0.0.1", help="Host to bind (use 0.0.0.0 for network access)")
    args = parser.parse_args()
    print(f"\n  Property Index Explorer running → http://{args.host}:{PORT}")
    print(f"  Excel  : {EXCEL_PATH}")
    print(f"  Graphs : {GRAPHS_DIR}")
    print(f"  Users  : {USERS_FILE}\n")
    app.run(host=args.host, debug=False, port=PORT)
